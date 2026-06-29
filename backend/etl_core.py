import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, NamedStyle, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import asyncio

class LegacyETL:
    def __init__(self, settings, report_date, dummy_code, import_files, export_path, ci_path, send_progress, handle_missing_categories, net_inv_mapping=None):
        self.settings = settings
        self.report_date = report_date
        self.dummy_code = dummy_code
        self.import_files = import_files
        self.export_path = export_path
        self.ci_path = ci_path
        self.send_progress = send_progress
        self.handle_missing_categories = handle_missing_categories
        self.net_inv_mapping = net_inv_mapping

    async def execute(self):
        report_date = self.report_date
        dummy_code = self.dummy_code
        export_path = self.export_path
        if not os.path.exists(export_path):
            os.makedirs(export_path, exist_ok=True)
            
        import_path = os.path.dirname(self.import_files['invoice'])
        ci_path = self.ci_path
        
        category_ref_path = self.settings.reference_path_category
        field_supervisors_df = pd.read_excel(self.settings.reference_path_fs)
        week_df = pd.read_excel(self.settings.reference_path_week)
        cdam_df = pd.read_excel(self.settings.reference_path_cdam)
        gt_channel_df = pd.read_excel(self.settings.reference_path_gt_channel)
        new_customer_df = pd.read_excel(self.settings.reference_path_new_customer)
        
        # Override file names to use the uploaded ones
        im_filename_inv = self.import_files['invoice']
        im_filename_ret = self.import_files['returns']
        im_filename_cust = self.import_files['customer']
        im_filename_route = self.import_files['route']
        im_filename_pl = self.import_files['price']
        im_filename_so = self.import_files['sales_order']
        
        category_df = pd.read_excel(category_ref_path)

        full_im_path_inv = im_filename_inv
        full_im_path_ret = im_filename_ret
        full_im_path_cust = im_filename_cust
        full_im_path_route = im_filename_route
        full_im_path_pl = im_filename_pl
        full_im_path_so = im_filename_so

        # Generate filenames based on the report date
        def generate_filename(report_type: str, date_obj: datetime, prefix: str = "KENEA") -> str:
            """Generate Excel filename with format: 'KENEA Report_Type_Month D, YYYY.xlsx'"""
            # Use '#' instead of '-' for Windows compatibility
            formatted_date = date_obj.strftime("%B %d, %Y").replace(" 0", " ")  # remove leading zero
            return f"{prefix} {report_type}_{formatted_date}.xlsx"
        
        # Build dynamic names from settings
        Net_Invoiced = generate_filename(self.settings.export_name_net_invoiced, report_date, prefix=self.settings.export_prefix)
        ex_filename_so = generate_filename(self.settings.export_name_sales_order, report_date, prefix=self.settings.export_prefix)
        ex_filename_ser_inv = generate_filename(self.settings.export_name_served_invoice, report_date, prefix=self.settings.export_prefix)
        ex_filename_cml = generate_filename(self.settings.export_name_cml, report_date, prefix=self.settings.export_prefix)
        

        
        await self.send_progress(f"Reading category reference data from {category_ref_path}...")
        # Load reference data
        category_df = pd.read_excel(category_ref_path)
        

        
        # Define import filenames
        im_filename_inv = "DMS-Invoice-on.xlsx"
        im_filename_ret = "DMS-Customer Returns-on.xlsx"
        im_filename_cust = "DMS-Customer-on.xlsx"
        im_filename_route = "DMS-Route-on.xlsx"
        im_filename_pl = "DMS-Price-on.xlsx"
        im_filename_so = "DMS-Sales Order-on.xlsx"
        
        # Define full import paths
        Net_Invoiced_path = os.path.join(export_path, Net_Invoiced)
        CML_path = os.path.join(export_path, ex_filename_cml)
        

        #Reading MONTHLY WRONG C.I. MONITORING
        ci_import_path = self.settings.reference_path_wrong_ci
        await self.send_progress(f"Reading Wrong C.I. Monitoring data from {ci_import_path}...")
        
        ci_df = pd.read_excel(ci_import_path, skiprows=4)
        
        ci_row_count = ci_df.count()
        
        await self.send_progress(f"Found {ci_row_count.iloc[0]} Recorded Wrong C.I. Deductions.")
        
        ci_checker_inv = ci_df.copy()
        ci_checker_inv = ci_checker_inv[['Invoice Filter']]
        
        ci_checker_ret = ci_df.copy()
        ci_checker_ret = ci_checker_ret[['Customer Return Filter']]
        
        
        #INVOICE ON Arrangement
        await self.send_progress(f"Extracting Invoice Data from {full_im_path_inv}...")
        
        df1 = pd.read_excel(full_im_path_inv)
        await self.send_progress(f"Invoice Data loaded successfully ({len(df1)} rows). Cleaning structures...")
        
        await self.send_progress("Filtering out Sales Tax items and applying C.I. deductions...")
        dfc1 = df1[['Invoice Date', 'Sold To Customer Number',
        'Product Code', 'Product/Item Description',
        'Total Item amount with Tax and Discount', 'Invoice Item Type','Invoice number']]
        inv = dfc1[dfc1['Invoice Item Type'] != 'ITM_SALES_TAX']
        
        if not ci_checker_inv.empty:
            ci_inv = inv.copy()
            ci_inv.insert(7,'Invoice Filter',0)
            ci_inv['Invoice Filter'] = ci_inv['Invoice number'].astype(str) + "_" + ci_inv['Product Code'].astype(str)
            ci_inv_processed = ci_inv[~ci_inv['Invoice Filter'].isin(ci_checker_inv['Invoice Filter'])]
            inv = ci_inv_processed.drop('Invoice Filter', axis=1)
        
        inv_f = inv.copy()
        inv_f = inv_f.drop('Invoice number', axis=1)
        inv_f.insert(6,'BO',0)
        inv_f.insert(7,'FG',0)
        inv_f['Total Item amount with Tax and Discount'] = inv_f['Total Item amount with Tax and Discount'].fillna(0)
        inv_f['Product Code'] = inv_f['Product Code'].astype(str)
        inv_f = inv_f.drop('Invoice Item Type', axis=1)
        #CUSTOMER RETURNS Arrangement
        await self.send_progress(f"Extracting Customer Returns Data from {full_im_path_ret}...")
        
        df2 = pd.read_excel(full_im_path_ret)
        await self.send_progress(f"Customer Returns Data loaded successfully ({len(df2)} rows). Processing deductions...")
        
        await self.send_progress("Pivoting Customer Returns by Facility Name (BO/FG)...")
        dfc2 = df2[['Customer Return Date', 'Sold To Customer Number', 'Product Code',
                    'Product Description', 'Facility Name', 'Estimated Product Return Amount','Customer Return Number']].copy()
        dfc2['Estimated Product Return Amount'] = dfc2['Estimated Product Return Amount'].fillna(0)
        ret_init = dfc2
        
        if not ci_checker_ret.empty:
            ci_ret = ret_init.copy()
            ci_ret.insert(7,'Customer Return Filter',0)
            ci_ret['Customer Return Filter'] = ci_ret['Customer Return Number'].astype(str) + "_" + ci_ret['Product Code'].astype(str)
            ci_ret_processed = ci_ret[~ci_ret['Customer Return Filter'].isin(ci_checker_ret['Customer Return Filter'])]
            ret_init = ci_ret_processed.drop('Customer Return Filter', axis=1)
        
        dfcr = ret_init.copy()
        dfcr = dfcr.drop('Customer Return Number', axis=1)
        dfcr.insert(4,'Total Item amount with Tax and Discount', 0)
        dfcr.insert(7,'with vat', 0)
        dfcr['with vat'] = dfcr['Estimated Product Return Amount'] * 1.12
        dfc3 = dfcr.copy()
        
        if dfc3['Facility Name'].str.contains('Virtual').any():
            dfc3['Facility Name'] = dfc3['Facility Name'].str.replace('Virtual', 'FG')
        
        # Extract only 'FG' or 'BO'
        dfc3['Facility Name'] = dfc3['Facility Name'].str.extract(r'\b(FG|BO)\b')
        
        cust_ret = dfc3.pivot_table(index=['Customer Return Date', 'Sold To Customer Number', 
        'Product Code', 'Product Description', 'Total Item amount with Tax and Discount'], columns='Facility Name', values='with vat', aggfunc=sum)
        
        ret_f = cust_ret.reset_index()
        
        ret_f['Product Code'] = ret_f['Product Code'].astype(str)
        if 'BO' in ret_f.columns and 'FG' not in ret_f.columns:
            ret_f['BO'] = ret_f['BO'].fillna(0)
            ret_f['FG'] = 0.0
        elif 'FG' in ret_f.columns and 'BO' not in ret_f.columns:
            ret_f['FG'] = ret_f['FG'].fillna(0)
            ret_f['BO'] = 0.0
        else:
            ret_f['BO'] = ret_f['BO'].fillna(0)
            ret_f['FG'] = ret_f['FG'].fillna(0)
        # SALES ORDERS Arrangements
        await self.send_progress(f"Extracting Sales Orders Data from {full_im_path_so}...")
        
        df3 = pd.read_excel(full_im_path_so)
        await self.send_progress(f"Sales Orders Data loaded successfully ({len(df3)} rows). Finding invoiced sales...")
        
        await self.send_progress("Filtering Invoiced Sales Orders and calculating VAT...")
        dfc3 = df3[['Last Modified Date', 'Sold To Customer number',
         'Product Code', 'Product Description', 'Total Product Amount', 'SO status', 'SO Number']]
        so_par = dfc3[dfc3['SO status'] == 'Invoiced']
        
        #Get Original Invoice
        dfc = df1[['Invoice Item Type','Product Code', 'SO Number','Invoice number']]
        inv_so = dfc[dfc['Invoice Item Type'] != 'ITM_SALES_TAX']
        
        if not ci_checker_inv.empty:
            inv_so = inv_so.copy()
            inv_so.insert(4,'Invoice Filter',0)
            inv_so.insert(5,'SO Filter',0)
        
            inv_so['Invoice Filter'] = inv_so['Invoice number'].astype(str) + "_" + inv_so['Product Code'].astype(str)
            inv_so['SO Filter'] = inv_so['SO Number'].astype(str) + "_" + inv_so['Product Code'].astype(str)
        
            inv_so_processed = inv_so[inv_so['Invoice Filter'].isin(ci_checker_inv['Invoice Filter'])]
            
            ci_so = inv_so_processed[['SO Filter']]
        
            inv_so = so_par.copy()
            inv_so.insert(7,'SO Filter',0)
            inv_so['SO Filter'] = inv_so['SO Number'].astype(str) + "_" + inv_so['Product Code'].astype(str)
            ci_so_processed = inv_so[~inv_so['SO Filter'].isin(ci_so['SO Filter'])]
            so_par = ci_so_processed.drop('SO Filter', axis=1)
        
        so_par = so_par.drop('SO Number', axis=1)
        so_par['Total Product Amount'] = so_par['Total Product Amount'].fillna(0)
        so_par.insert(6,'with vat', 0)
        so_par['with vat'] = so_par['Total Product Amount'] * 1.12
        
        so_f = so_par.copy()
        so_f = so_f.drop(['Total Product Amount', 'SO status'], axis=1)
        #Customer Master List Reading
        await self.send_progress(f"Reading Customer Master List from {full_im_path_cust}...")
        
        dms_cust_df = pd.read_excel(full_im_path_cust)
        await self.send_progress(f"Customer Master List loaded successfully ({len(dms_cust_df)} customers). Organizing CML data structures...")
        
        await self.send_progress("Organizing Customer and CML Data structures...")
        cust_df = dms_cust_df.copy()
        cml_df = dms_cust_df.copy()
        
        cust_df = cust_df[[
            'NEXT_UP_NUMBER',
            'CUSTOMER_NAME',
            'PARTY_CLASSIFICATION_DESCRIPTION',
            'KEY_ACCOUNT',
            'SALES_REP_ID',
            'SALES_REP_NAME',
            'BRANCH_NAME',
            'GEO_LOCATION_HIERARCHYDESCRIPTION',
            'CITY',
            'STATE_PROVINCE',
            'CHANNEL'
        ]]
        
        #Customer Master List Arrangement and Finalization
        
        cml_df = cml_df[[
            'BRANCH_NAME',
            'CHANNEL',
            'SALES_REP_ID',
            'SALES_REP_NAME',
            'NEXT_UP_NUMBER',
            'CUSTOMER_NAME',
            'GEO_LOCATION_HIERARCHYDESCRIPTION',
            'CITY',
            'STATE_PROVINCE',
            'STATUS',
            'RETAIL_ENVIRONMENT',
            'PARTY_CLASSIFICATION_DESCRIPTION',
            'CREATED_DATE'
        ]]
        
        # New Customer List Arrangement
        new_customer_df = new_customer_df[[
            'NEXT_UP_NUMBER',
            'NEW_CUSTOMER'
        ]]
        
        #Sales Force Reading
        await self.send_progress(f"Extracting Route Data from {full_im_path_route}...")
        
        dms_route_df = pd.read_excel(full_im_path_route)
        await self.send_progress(f"Route Data loaded successfully ({len(dms_route_df)} routes). Mapping coverage to CML...")
        
        await self.send_progress("Merging Route Data with CML...")
        dms_route = dms_route_df.copy()[[
            'customer_no',
            'route_code'
        ]]
        
        # Clean up CDAM DF
        # Set proper column names from row 1 and skip the first 2 rows
        await self.send_progress("Cleaning up CDAM structure...")
        cdam_df_clean = cdam_df.iloc[2:].copy()
        cdam_df_clean.columns = ['PARTY_CLASSIFICATION_DESCRIPTION', 'CDAM']
        cdam_df_clean = cdam_df_clean.reset_index(drop=True)
        cdam_df_clean = cdam_df_clean.dropna()
        
        # Clean up FS DF
        # Field supervisors already has correct column names from Excel; no need to skip rows
        await self.send_progress("Applying Field Supervisor names and roles...")
        fs_df_clean = field_supervisors_df.copy()
        fs_df_clean = fs_df_clean.reset_index(drop=True)
        fs_df_clean = fs_df_clean.dropna()
        
        # Parse route_code to extract COVERAGE DAY, WKLY COVERAGE, and FREQ
        def parse_route_code(route_code):
            """
            Parse route_code to extract COVERAGE_DAY, WKLY_COVERAGE, and FREQ
            Examples:
            - KEA0001_TUE_WKLY → ('TUE', 'WKLY', 'F4')
            - KEA0002_MON_WK1_WK3 → ('MON', 'W1&W3', 'F2')
            - KEA0003_WED_WK2_WK4 → ('WED', 'W2&W4', 'F2')
            """
            parts = route_code.split('_')
            
            # Second part is the day (TUE, MON, WED, etc.)
            coverage_day = parts[1] if len(parts) > 1 else ''
            
            # Third part onwards is the weekly coverage (WKLY or WK1_WK3, WK2_WK4, etc.)
            wkly_coverage = '_'.join(parts[2:]) if len(parts) > 2 else ''
            
            # Convert specific weekly coverage patterns
            if wkly_coverage == 'WK1_WK3':
                wkly_coverage = 'W1&W3'
            elif wkly_coverage == 'WK2_WK4':
                wkly_coverage = 'W2&W4'
            
            # Determine frequency based on wkly_coverage
            if 'WKLY' in wkly_coverage:
                freq = 'F4'
            elif 'WK' in wkly_coverage or '&' in wkly_coverage:
                freq = 'F2'
            else:
                freq = ''
            
            return coverage_day, wkly_coverage, freq
        
        # Apply the parsing function
        dms_route_df[['COVERAGE_DAY', 'WKLY_COVERAGE', 'FREQ']] = dms_route_df['route_code'].apply(
            lambda x: pd.Series(parse_route_code(x))
        )
            
        # Merge dms_route_df with cml_df on customer_no and add routing information
        # First, prepare dms_route_df by renaming the customer_no column
        dms_route_merge = dms_route_df[['customer_no', 'COVERAGE_DAY', 'WKLY_COVERAGE', 'FREQ']].copy()
        dms_route_merge = dms_route_merge.rename(columns={'customer_no': 'NEXT_UP_NUMBER'})
        
        # Merge dms_route_df with cml_df on NEXT_UP_NUMBER
        cml_df = cml_df.merge(dms_route_merge, on='NEXT_UP_NUMBER', how='left')
        
        # Set COVERAGE_DAY, WKLY_COVERAGE, FREQ to "NA" if STATUS is "Blocked/On hold"
        cml_df.loc[cml_df['STATUS'] == 'Blocked/On hold', 'COVERAGE_DAY'] = 'NA'
        cml_df.loc[cml_df['STATUS'] == 'Blocked/On hold', 'WKLY_COVERAGE'] = 'NA'
        cml_df.loc[cml_df['STATUS'] == 'Blocked/On hold', 'FREQ'] = 'NA'
        
        # Add FREQ_COUNT column based on STATUS
        cml_df['FREQ_COUNT'] = 0  # Default value
        cml_df.loc[cml_df['STATUS'] == 'Active/Approved', 'FREQ_COUNT'] = 1
        cml_df.loc[cml_df['STATUS'] == 'Blocked/On hold', 'FREQ_COUNT'] = 0
        
        # Delete rows where STATUS is "Approval On Hold"
        cml_df = cml_df[cml_df['STATUS'] != 'Approval On Hold'].reset_index(drop=True)
        
        # Merge cdam_df_clean based on PARTY_CLASSIFICATION_DESCRIPTION
        cml_df = cml_df.merge(cdam_df_clean, on='PARTY_CLASSIFICATION_DESCRIPTION', how='left')
        
        # CUSTOMIZE CDAM FOR SPECIFIC CUSTOMERS
        # Set CDAM to "H ACCT" for specific customer codes
        special_customers = ['MKXCH00001', 'MKXFB00001', 'KNEC100001', 'KEAC100001']
        cml_df.loc[cml_df['NEXT_UP_NUMBER'].isin(special_customers), 'CDAM'] = 'H ACCT'
        
        # Merge fs_df_clean based on SALES_REP_ID
        cml_df = cml_df.merge(fs_df_clean, on='SALES_REP_ID', how='left')
        
        # Merge new_customer_df based on NEXT_UP_NUMBER
        cml_df = cml_df.merge(new_customer_df, on='NEXT_UP_NUMBER', how='left')
        # Robustly treat empty/whitespace or missing NEW_CUSTOMER as 'YES'
        cml_df['NEW_CUSTOMER'] = cml_df['NEW_CUSTOMER'].replace(r'^\s*$', pd.NA, regex=True).fillna('YES')
        
        # Create the final dataframe with selected columns and rename them
        cml_final_df = cml_df[[
            'BRANCH_NAME',
            'CDAM',
            'FS',
            'CHANNEL',
            'SALES_REP_ID',
            'SALES_REP_NAME',
            'NEXT_UP_NUMBER',
            'CUSTOMER_NAME',
            'GEO_LOCATION_HIERARCHYDESCRIPTION',
            'CITY',
            'STATE_PROVINCE',
            'STATUS',
            'RETAIL_ENVIRONMENT',
            'PARTY_CLASSIFICATION_DESCRIPTION',
            'COVERAGE_DAY',
            'WKLY_COVERAGE',
            'FREQ_COUNT',
            'FREQ',
            'CREATED_DATE',
            'NEW_CUSTOMER'
        ]].copy()
        
        # Remove duplicates based on NEXT_UP_NUMBER, keeping the most recent CREATED_DATE
        await self.send_progress("Removing CML duplicates based on creation date...")
        cml_final_df = cml_final_df.sort_values('CREATED_DATE', ascending=True).drop_duplicates(subset='NEXT_UP_NUMBER', keep='first')
        
        # Convert CREATED_DATE to datetime for proper sorting, then sort by CREATED_DATE and NEXT_UP_NUMBER
        cml_final_df['CREATED_DATE_DT'] = pd.to_datetime(cml_final_df['CREATED_DATE'], format='%m/%d/%Y')
        cml_final_df = cml_final_df.sort_values(['CREATED_DATE_DT', 'NEXT_UP_NUMBER'])
        cml_final_df = cml_final_df.drop('CREATED_DATE_DT', axis=1)
        
        # Rename columns
        cml_final_df = cml_final_df.rename(columns={
            'BRANCH_NAME': 'BRANCH NAME',
            'SALES_REP_ID': 'SALES REP ID',
            'SALES_REP_NAME': 'SALES REP NAME',
            'NEXT_UP_NUMBER': 'CUSTOMER CODE',
            'CUSTOMER_NAME': 'CUSTOMER NAME',
            'GEO_LOCATION_HIERARCHYDESCRIPTION': 'BARANGAY',
            'STATE_PROVINCE': 'PROVINCE',
            'RETAIL_ENVIRONMENT': 'RETAIL ENVIRONMENT',
            'PARTY_CLASSIFICATION_DESCRIPTION': 'PARTY CLASSIFICATION DESCRIPTION',
            'COVERAGE_DAY': 'COVERAGE DAY',
            'WKLY_COVERAGE': 'WKLY COVERAGE',
            'FREQ_COUNT': 'FREQ COUNT',
            'NEW_CUSTOMER': 'NEW CUSTOMER'
        })
        
        # Remove CREATED_DATE column before final output
        cml_final_df = cml_final_df.drop('CREATED_DATE', axis=1)
        
        # Arrange columns in the specified order
        cml_final_df = cml_final_df[[
            'BRANCH NAME',
            'CDAM',
            'FS',
            'CHANNEL',
            'SALES REP ID',
            'SALES REP NAME',
            'CUSTOMER CODE',
            'CUSTOMER NAME',
            'BARANGAY',
            'CITY',
            'PROVINCE',
            'STATUS',
            'RETAIL ENVIRONMENT',
            'PARTY CLASSIFICATION DESCRIPTION',
            'COVERAGE DAY',
            'WKLY COVERAGE',
            'FREQ COUNT',
            'FREQ',
            'NEW CUSTOMER'
        ]]
        
        # CML Finalization
        
        # Step 1: Export to Excel with a blank row at the top
        
        temp_output = 'temp_file_output.xlsx'
        temp_file = os.path.join(import_path, temp_output)
        
        with pd.ExcelWriter(temp_file, engine='openpyxl') as writer:
            # Write to row 2 (startrow=1) to leave row 1 blank
            cml_final_df.to_excel(writer, index=False, startrow=3)
        
        # Step 2: Load workbook and apply formatting
        wb = load_workbook(temp_file)
        ws = wb.active
        
        # Step 3: Style header (row 4)
        header_fill = PatternFill(start_color='44546A', end_color='44546A', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        header_border = Border(left=Side(style=None), right=Side(style=None), top=Side(style=None), bottom=Side(style=None))
        
        # Set row height for header (row 4)
        ws.row_dimensions[4].height = 45
        
        for cell in ws[4]:  # Header row
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = header_border
        
        # Add SUBTOTAL formulas in row 3 (above header) with red font
        subtotal_font = Font(bold=True, color='FF0000')
        subtotal_alignment = Alignment(horizontal='center')
        
        # Get the last row number (data starts at row 5, so last row is 5 + len(cml_final_df) - 1)
        last_row = 4 + len(cml_final_df)
        
        for col_num in range(1, len(cml_final_df.columns) + 1):
            col_letter = get_column_letter(col_num)
            subtotal_cell = ws.cell(row=3, column=col_num)
            subtotal_cell.value = f"=SUBTOTAL(103,{col_letter}5:{col_letter}{last_row})"
            subtotal_cell.font = subtotal_font
            subtotal_cell.alignment = subtotal_alignment
        
        # Auto-fit all columns (equivalent to double-clicking column headers in Excel)
        for col_num in range(1, len(cml_final_df.columns) + 1):
            col_letter = get_column_letter(col_num)
            ws.column_dimensions[col_letter].auto_size = True
        
        
        # Save changes
        ws.sheet_view.showGridLines = False
        wb.properties.author = "Joshua Ocampo"
        wb.save(CML_path)
        await self.send_progress("ETL Data processing completed successfully! All files have been exported.")
        #PriceList Reading
        await self.send_progress(f"Extracting Price List Data from {full_im_path_pl}...")
        
        dms_pl_df = pd.read_excel(full_im_path_pl)
        
        await self.send_progress("Processing Standard Price configurations...")
        dms_pl = dms_pl_df.copy()
        
        pl_df2 = dms_pl
        
        pl_df2 = pl_df2[['product_code', 'product_description', 'uom_description', 'selling_price', 'cust_class', 'cust_channel']]
        pl_df2.insert(4,'with vat', 0)
        pl_df2['with vat'] = pl_df2['selling_price'] * 1.12
        pl_df3 = pl_df2[pl_df2['cust_class'] != 'BEV Dealer']
        
        pl_df_m0 = pl_df3
        pl_df_m2 = pl_df3
        
        pl_df_m0 = pl_df_m0[pl_df_m0['cust_channel'] != 'VAN(EXTRUCK)']
        pl_df_m2 = pl_df_m2[pl_df_m2['cust_channel'] == 'VAN(EXTRUCK)']
        
        #Price List M0
        
        pl_df_m0_f = pl_df_m0.pivot_table(index=['product_code', 'product_description',], columns='uom_description', values='with vat', aggfunc=sum)
        
        pl_df_m0_f_1 = pl_df_m0_f.reset_index()
        
        pl_df_m0_f_1['Case'] = pl_df_m0_f_1['Case'].fillna(0)
        pl_df_m0_f_1['Piece'] = pl_df_m0_f_1['Piece'].fillna(0)
        pl_df_m0_f_1['Subcase'] = pl_df_m0_f_1['Subcase'].fillna(0)
        
        pl_m0_final = pl_df_m0_f_1[['product_code', 'product_description', 'Case', 'Subcase', 'Piece']]
        pl_m0_final['product_code'] = pl_m0_final['product_code'].astype(str)
        
        ex_filename_pl_m0 = "M0 Pricelist.xlsx"
        full_path_pl_m0 = os.path.join(export_path, ex_filename_pl_m0)
        
        export_pl_m0 = pl_m0_final
        export_pl_m0.to_excel(full_path_pl_m0, index=False)
        
        # Set author metadata
        wb_m0 = load_workbook(full_path_pl_m0)
        wb_m0.properties.author = "Joshua Ocampo"
        wb_m0.save(full_path_pl_m0)
        
        #Price List M2
        
        pl_df_m2_f = pl_df_m2.pivot_table(index=['product_code', 'product_description',], columns='uom_description', values='with vat', aggfunc=sum)
        
        pl_df_m2_f_1 = pl_df_m2_f.reset_index()
        
        pl_df_m2_f_1['Case'] = pl_df_m2_f_1['Case'].fillna(0)
        pl_df_m2_f_1['Piece'] = pl_df_m2_f_1['Piece'].fillna(0)
        pl_df_m2_f_1['Subcase'] = pl_df_m2_f_1['Subcase'].fillna(0)
        
        pl_m2_final = pl_df_m2_f_1[['product_code', 'product_description', 'Case', 'Subcase', 'Piece']]
        pl_m2_final['product_code'] = pl_m2_final['product_code'].astype(str)
        
        ex_filename_pl_m2 = "M3 Pricelist.xlsx"
        full_path_pl_m2 = os.path.join(export_path, ex_filename_pl_m2)
        
        export_pl_m2 = pl_m2_final
        export_pl_m2.to_excel(full_path_pl_m2, index=False)
        
        # Set author metadata
        wb_m2 = load_workbook(full_path_pl_m2)
        wb_m2.properties.author = "Joshua Ocampo"
        wb_m2.save(full_path_pl_m2)
        
        # Get Volume - Price Arrangements
        pl_m0_final_reference = pl_m0_final[['product_code', 'Case']].copy()
        pl_m2_final_reference = pl_m2_final[['product_code', 'Case']].copy()
        # Function to edit missing categories via GUI
        
        # Prepare data for category checking
        cat_ck = category_df.copy()
        inv_ck = inv_f[['Product Code', 'Product/Item Description']].rename(
            columns={'Product Code': 'SKU CODE', 'Product/Item Description': 'SKU NAME'}
        ).drop_duplicates(subset=['SKU CODE']).reset_index(drop=True)
        
        ret_ck = ret_f[['Product Code', 'Product Description']].rename(
            columns={'Product Code': 'SKU CODE', 'Product Description': 'SKU NAME'}
        ).drop_duplicates(subset=['SKU CODE']).reset_index(drop=True)
        
        so_ck = so_f[['Product Code', 'Product Description']].rename(
            columns={'Product Code': 'SKU CODE', 'Product Description': 'SKU NAME'}
        ).drop_duplicates(subset=['SKU CODE']).reset_index(drop=True)
        
        # do NOT pre-merge the category reference into all_products here — keep SKU NAME coming from source tables
        all_products = pd.concat([inv_ck, ret_ck, so_ck], ignore_index=True)
        all_products = all_products.drop_duplicates(subset=['SKU CODE']).reset_index(drop=True)
        
        # Call editor and reload saved reference
        category_df = await self.handle_missing_categories(category_df, all_products, category_ref_path)
        
        category_df = pd.read_excel(category_ref_path, dtype={'SKU CODE': str})
        #COMBINE INVOICED AND RETURNS to NET INVOICED
        
        inv_f = inv_f.copy()
        inv_f.rename(columns={
        'Invoice Date': 'DATE',
        'Sold To Customer Number': 'ACCOUNT CODE',
        # 'Sold To Customer Name': 'ACCOUNT NAME',
        'Product Code': 'SKU CODE',
        'Product/Item Description': 'SKU NAME',
        'Total Item amount with Tax and Discount': 'SERVED INVOICE',
        'BO': 'BAD RETURNS',
        'FG': 'GOOD RETURNS'
        }, inplace=True)
        
        ret_f = ret_f.copy()
        ret_f.rename(columns={
        'Customer Return Date': 'DATE',
        'Sold To Customer Number': 'ACCOUNT CODE',
        'Product Code': 'SKU CODE', 
        'Product Description': 'SKU NAME',
        'Total Item amount with Tax and Discount': 'SERVED INVOICE',
        'BO': 'BAD RETURNS',
        'FG': 'GOOD RETURNS'
        }, inplace=True)
        
        net_inv = pd.concat([inv_f, ret_f], axis=0, ignore_index=True)
        
        # Customer Details Arrangements
        cml_df = cust_df.rename(columns={
            'NEXT_UP_NUMBER': 'ACCOUNT CODE'
        })
        
        net_inv_f = net_inv.merge(cml_df, on='ACCOUNT CODE', how='left')
        
        # Drop SKU NAME from category_df if it exists to avoid confusion in merges
        if 'SKU NAME' in category_df.columns:
            category_df = category_df.drop(['SKU NAME'], axis=1)
        net_inv_f_l1 = net_inv_f.merge(category_df,  on='SKU CODE', how='left')
        #CML Addons
        await self.send_progress("Appending Field Supervisors, Week, CDAM, and GT Channel logic to CML...")
        
        net_inv_f_l2 = net_inv_f_l1.merge(field_supervisors_df, on='SALES_REP_ID', how='left')
        
        net_inv_f_l3 = net_inv_f_l2.merge(week_df, on='DATE', how='inner')
        
        pl_m0_final_reference = pl_m0_final_reference.rename(columns={
            'product_code': 'SKU CODE',
            'Case': 'SKU PRICE REFERENCE_M0'
        })
        
        pl_m2_final_reference = pl_m2_final_reference.rename(columns={
            'product_code': 'SKU CODE',
            'Case': 'SKU PRICE REFERENCE_M2'
        })
        
        price_reference = pl_m0_final_reference.merge(pl_m2_final_reference, on='SKU CODE', how='outer')
        
        # Clean up SKU CODE in net_inv_f_l3 to ensure it matches the format in price_reference
        net_inv_f_l3['SKU CODE'] = net_inv_f_l3['SKU CODE'].str.replace('_old', '', regex=False)
        
        # Merge price reference into net invoiced data
        net_inv_f_l4 = net_inv_f_l3.merge(price_reference, on='SKU CODE', how='left')
        
        net_inv_f_l4.insert(5,'VALUE', 0)
        net_inv_f_l4['VALUE'] = net_inv_f_l4['SERVED INVOICE'] - net_inv_f_l4['BAD RETURNS'] - net_inv_f_l4['GOOD RETURNS']
        
        # ensure numeric and avoid divide-by-zero when computing VOLUME
        net_inv_f_l4['SKU PRICE REFERENCE_M0'] = pd.to_numeric(net_inv_f_l4['SKU PRICE REFERENCE_M0'], errors='coerce').fillna(0)
        net_inv_f_l4['SKU PRICE REFERENCE_M2'] = pd.to_numeric(net_inv_f_l4['SKU PRICE REFERENCE_M2'], errors='coerce').fillna(0)
        
        net_inv_f_l4['VOLUME'] = 0.0
        van_mask = net_inv_f_l4['CHANNEL'] == 'VAN(EXTRUCK)'
        book_mask = net_inv_f_l4['CHANNEL'] == 'BOOK(Booking)'
        
        # Use the proper price reference for each channel
        van_price_mask = van_mask & (net_inv_f_l4['SKU PRICE REFERENCE_M2'] != 0)
        net_inv_f_l4.loc[van_price_mask, 'VOLUME'] = (
            net_inv_f_l4.loc[van_price_mask, 'VALUE'] /
            net_inv_f_l4.loc[van_price_mask, 'SKU PRICE REFERENCE_M2']
        )
        
        book_price_mask = book_mask & (net_inv_f_l4['SKU PRICE REFERENCE_M0'] != 0)
        net_inv_f_l4.loc[book_price_mask, 'VOLUME'] = (
            net_inv_f_l4.loc[book_price_mask, 'VALUE'] /
            net_inv_f_l4.loc[book_price_mask, 'SKU PRICE REFERENCE_M0']
        )
        
        # Merge GT Channel based on PARTY_CLASSIFICATION_DESCRIPTION
        gt_channel = gt_channel_df.copy()
        gt_channel = gt_channel[['PARTY_CLASSIFICATION_DESCRIPTION', 'GT_Channel']]
        net_inv_f_l5 = net_inv_f_l4.merge(gt_channel, on='PARTY_CLASSIFICATION_DESCRIPTION', how='left')
        
        # Post-Net Invoiced
        export_net_inv = net_inv_f_l5
        export_net_inv['RD NAME'] = 'Kimberlin'
        
        export_net_inv = export_net_inv.sort_values(by=['DATE', 'ACCOUNT CODE'])
        
        if not self.net_inv_mapping:
            raise Exception("Activation mapping missing! Unauthorized access.")
            
        export_net_inv.rename(columns=self.net_inv_mapping, inplace=True)
        
        export_net_inv_final = export_net_inv[[
            'RD Name',
            'Date',
            'Week',
            'Branch Name',
            'Employee Code',
            'Employee Name',
            'Channel',
            'Sold To Customer number',
            'Sold To Customer Name',
            'Category',
            'Product Code',
            'Product Description',
            'Volume',
            'Net Value',
            'Good Stock Returns',
            'Bad Stock Returns',
            'Channel_Classification',
            'Brgy',
            'Town',
            'Province',
            'FS',
            'RTM Model',
            'GT Channel'
        ]]
        
        # Net Invoiced Finalization
        
        # Step 1: Export to Excel with a blank row at the top
        
        temp_output = 'temp_file_output.xlsx'
        temp_file = os.path.join(import_path, temp_output)
        
        with pd.ExcelWriter(temp_file, engine='openpyxl') as writer:
            # Write to row 2 (startrow=1) to leave row 1 blank
            export_net_inv_final.to_excel(writer, index=False, startrow=1)
        
        # Step 2: Load workbook and apply formatting
        wb = load_workbook(temp_file)
        ws = wb.active
        
        # Step 3: Style header (row 2)
        header_fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        header_alignment = Alignment(horizontal='center')
        
        for cell in ws[2]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Step 4: Format numeric columns with comma style
        accounting_style = NamedStyle(name="accounting_style", number_format='_(* #,##0.00_);_(* (#,##0.00);_(* "-"_);_(@_)')
        for col in range(13, 17):  # Assuming numeric columns start from column 2
            for row in range(3, ws.max_row + 1):  # Data starts from row 3
                cell = ws.cell(row=row, column=col)
                if isinstance(cell.value, (int, float)):
                    cell.style = accounting_style
        
        sum_row = ws.max_row + 1
        start_range = 3
        end_range = ws.max_row
        
        for col in range(13, 17):
            col_letter = get_column_letter(col)
            sum_cell = ws.cell(row=sum_row, column=col)
            sum_cell.value = f"=SUM({col_letter}{start_range}:{col_letter}{end_range})"
            sum_cell.font = Font(bold=True, color="FF0000")
            sum_cell.number_format = '_(* #,##0.00_);_(* (#,##0.00);_(* "-"_);_(@_)'
        
        # List the column letters or indexes you want to auto-fit
        columns_to_adjust = ['M', 'N', 'O', 'P']  # or use [2, 4] for indexes
        
        for col_id in columns_to_adjust:
            # Convert index to letter if needed
            col_letter = get_column_letter(col_id) if isinstance(col_id, int) else col_id
            max_length = 0
        
            for cell in ws[col_letter]:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
        
            ws.column_dimensions[col_letter].width = max_length + 2  # Add padding
        
        
        # Save changes
        ws.sheet_view.showGridLines = False
        wb.properties.author = "Joshua Ocampo"
        wb.save(Net_Invoiced_path)
        # SALES ORDER Final Arrangements and Export
        
        cml_df = cust_df.rename(columns={
            'NEXT_UP_NUMBER': 'Sold To Customer number'
        })
        cat_df = category_df.rename(columns={
            'SKU CODE': 'Product Code'
        })
        wk_df = week_df.rename(columns={
            'DATE': 'Last Modified Date'
        })
        pl_df_m0 = pl_m0_final_reference.rename(columns={
            'SKU CODE': 'Product Code'
        })
        pl_df_m2 = pl_m2_final_reference.rename(columns={
            'SKU CODE': 'Product Code'
        })
        
        sales_orders_df1 = so_f.merge(cml_df, on='Sold To Customer number', how='left')
        
        sales_orders_df2 = sales_orders_df1.merge(cat_df,  on='Product Code', how='left')
        
        sales_orders_df3 = sales_orders_df2.merge(field_supervisors_df, on='SALES_REP_ID', how='left')
        
        sales_orders_df4 = sales_orders_df3.merge(wk_df, on='Last Modified Date', how='inner')
        
        sales_orders_df4['Product Code'] = sales_orders_df4['Product Code'].str.replace('_old', '', regex=False)
        
        sales_orders_df5 = sales_orders_df4.merge(pl_df_m0, on='Product Code', how='left')
        sales_orders_df5 = sales_orders_df5.merge(pl_df_m2, on='Product Code', how='left')
        
        sales_orders_df5['SKU PRICE REFERENCE_M0'] = pd.to_numeric(sales_orders_df5['SKU PRICE REFERENCE_M0'], errors='coerce').fillna(0)
        sales_orders_df5['SKU PRICE REFERENCE_M2'] = pd.to_numeric(sales_orders_df5['SKU PRICE REFERENCE_M2'], errors='coerce').fillna(0)
        
        sales_orders_df5['VOLUME'] = 0.0
        van_mask = sales_orders_df5['CHANNEL'] == 'VAN(EXTRUCK)'
        book_mask = sales_orders_df5['CHANNEL'] == 'BOOK(Booking)'
        
        van_price_mask = van_mask & (sales_orders_df5['SKU PRICE REFERENCE_M2'] != 0)
        sales_orders_df5.loc[van_price_mask, 'VOLUME'] = (
            sales_orders_df5.loc[van_price_mask, 'with vat'] /
            sales_orders_df5.loc[van_price_mask, 'SKU PRICE REFERENCE_M2']
        )
        
        book_price_mask = book_mask & (sales_orders_df5['SKU PRICE REFERENCE_M0'] != 0)
        sales_orders_df5.loc[book_price_mask, 'VOLUME'] = (
            sales_orders_df5.loc[book_price_mask, 'with vat'] /
            sales_orders_df5.loc[book_price_mask, 'SKU PRICE REFERENCE_M0']
        )
        
        sales_orders_df5['RD NAME'] = 'Kimberlin'
        
        export_sales_order = sales_orders_df5.copy()
        
        export_sales_order.rename(columns={
            'RD NAME': 'RD Name',
            'Last Modified Date': 'SO Date',
            'WEEK': 'Week',
            'BRANCH_NAME': 'Branch Name',
            'SALES_REP_ID': 'Employee Code',
            'SALES_REP_NAME': 'Employee Name',
            'KEY_ACCOUNT': 'Channel',
            'Sold To Customer number': 'Sold To Customer Number',
            'CUSTOMER_NAME': 'Sold To Customer Name',
            'CATEGORY': 'Category',
            'VOLUME': 'Volume',
            'with vat': 'Value',
            'PARTY_CLASSIFICATION_DESCRIPTION': 'Channel Type'
        }, inplace=True)
        
        export_sales_order_final = export_sales_order[[
            'RD Name',
            'SO Date',
            'Week',
            'Branch Name',
            'Employee Code',
            'Employee Name',
            'Channel',
            'Sold To Customer Number',
            'Sold To Customer Name',
            'Category',
            'Product Code',
            'Product Description',
            'Volume',
            'Value',
            'FS',
            'Channel Type'
        ]].copy()
        
        so_df_export = export_sales_order_final
        so_df_export_final = so_df_export.sort_values(by=['SO Date', 'Sold To Customer Number'])
        
        # SERVED INVOICE Finalization
        
        # Step 1: Export to Excel with a blank row at the top
        temp_output = 'temp_file_output.xlsx'
        temp_file = os.path.join(import_path, temp_output)
        
        with pd.ExcelWriter(temp_file, engine='openpyxl') as writer:
            # Write to row 2 (startrow=1) to leave row 1 blank
            so_df_export_final.to_excel(writer, index=False, startrow=1)
        
        # Step 2: Load workbook and apply formatting
        wb = load_workbook(temp_file)
        ws = wb.active
        
        # Step 3: Style header (row 2)
        header_fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        header_alignment = Alignment(horizontal='center')
        
        for cell in ws[2]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Step 4: Format numeric columns with comma style
        accounting_style = NamedStyle(name="accounting_style", number_format='_(* #,##0.00_);_(* (#,##0.00);_(* "-"_);_(@_)')
        for col in range(13, 15):  # Assuming numeric columns start from column 2
            for row in range(3, ws.max_row + 1):  # Data starts from row 3
                cell = ws.cell(row=row, column=col)
                if isinstance(cell.value, (int, float)):
                    cell.style = accounting_style
        
        sum_row = ws.max_row + 1
        start_range = 3
        end_range = ws.max_row
        
        for col in range(13, 15):
            col_letter = get_column_letter(col)
            sum_cell = ws.cell(row=sum_row, column=col)
            sum_cell.value = f"=SUM({col_letter}{start_range}:{col_letter}{end_range})"
            sum_cell.font = Font(bold=True, color="FF0000")
            sum_cell.number_format = '_(* #,##0.00_);_(* (#,##0.00);_(* "-"_);_(@_)'
        
        # List the column letters or indexes you want to auto-fit
        columns_to_adjust = ['M', 'N']  # or use [2, 4] for indexes
        
        for col_id in columns_to_adjust:
            # Convert index to letter if needed
            col_letter = get_column_letter(col_id) if isinstance(col_id, int) else col_id
            max_length = 0
        
            for cell in ws[col_letter]:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
        
            ws.column_dimensions[col_letter].width = max_length + 2  # Add padding
        
        # Save changes
        full_path_so = os.path.join(export_path, ex_filename_so)
        
        ws.sheet_view.showGridLines = False
        wb.properties.author = "Joshua Ocampo"
        wb.save(full_path_so)
        # SERVED INVOICE Arrangements
        
        ser_inv = inv_f.copy()
        ser_inv_df = ser_inv.drop(['BAD RETURNS', 'GOOD RETURNS'], axis=1)
        
        cml_df = cust_df.rename(columns={
            'NEXT_UP_NUMBER': 'ACCOUNT CODE'
        })
        
        ser_inv_df1 = ser_inv_df.merge(cml_df, on='ACCOUNT CODE', how='left')
        
        ser_inv_f_l1 = ser_inv_df1.merge(category_df,  on='SKU CODE', how='left')
        
        ser_inv_f_l2 = ser_inv_f_l1.merge(field_supervisors_df, on='SALES_REP_ID', how='left')
        
        ser_inv_f_l3 = ser_inv_f_l2.merge(week_df, on='DATE', how='inner')
        
        ser_inv_f_l3['SKU CODE'] = ser_inv_f_l3['SKU CODE'].str.replace('_old', '', regex=False)
        
        ser_inv_f_l4 = ser_inv_f_l3.merge(pl_m0_final_reference, on='SKU CODE', how='left')
        ser_inv_f_l4 = ser_inv_f_l4.merge(pl_m2_final_reference, on='SKU CODE', how='left')
        
        ser_inv_f_l4['SKU PRICE REFERENCE_M0'] = pd.to_numeric(ser_inv_f_l4['SKU PRICE REFERENCE_M0'], errors='coerce').fillna(0)
        ser_inv_f_l4['SKU PRICE REFERENCE_M2'] = pd.to_numeric(ser_inv_f_l4['SKU PRICE REFERENCE_M2'], errors='coerce').fillna(0)
        
        ser_inv_f_l4['VOLUME'] = 0.0
        van_mask = ser_inv_f_l4['CHANNEL'] == 'VAN(EXTRUCK)'
        book_mask = ser_inv_f_l4['CHANNEL'] == 'BOOK(Booking)'
        
        van_price_mask = van_mask & (ser_inv_f_l4['SKU PRICE REFERENCE_M2'] != 0)
        ser_inv_f_l4.loc[van_price_mask, 'VOLUME'] = (
            ser_inv_f_l4.loc[van_price_mask, 'SERVED INVOICE'] /
            ser_inv_f_l4.loc[van_price_mask, 'SKU PRICE REFERENCE_M2']
        )
        
        book_price_mask = book_mask & (ser_inv_f_l4['SKU PRICE REFERENCE_M0'] != 0)
        ser_inv_f_l4.loc[book_price_mask, 'VOLUME'] = (
            ser_inv_f_l4.loc[book_price_mask, 'SERVED INVOICE'] /
            ser_inv_f_l4.loc[book_price_mask, 'SKU PRICE REFERENCE_M0']
        )
        
        ser_inv_f_l4['RD NAME'] = 'Kimberlin'
        
        
        export_ser_inv = ser_inv_f_l4.copy()
        
        export_ser_inv.rename(columns={
            'RD NAME': 'RD Name',
            'DATE': 'Invoice Date',
            'WEEK': 'Week',
            'BRANCH_NAME': 'Branch Name',
            'SALES_REP_ID': 'Employee Code',
            'SALES_REP_NAME': 'Employee Name',
            'KEY_ACCOUNT': 'Channel',
            'ACCOUNT CODE': 'Sold To Customer Number',
            'CUSTOMER_NAME': 'Sold To Customer Name',
            'CATEGORY': 'Category',
            'SKU CODE': 'Product Code',
            'SKU NAME': 'Product Description',
            'VOLUME': 'Volume',
            'SERVED INVOICE': 'Value',
            'PARTY_CLASSIFICATION_DESCRIPTION': 'Channel Type'
        }, inplace=True)
        
        export_ser_inv_final = export_ser_inv[[
            'RD Name',
            'Invoice Date',
            'Week',
            'Branch Name',
            'Employee Code',
            'Employee Name',
            'Channel',
            'Sold To Customer Number',
            'Sold To Customer Name',
            'Category',
            'Product Code',
            'Product Description',
            'Volume',
            'Value',
            'FS',
            'Channel Type'
        ]].copy()
        
        ser_inv_df_export = export_ser_inv_final
        ser_inv_df_export_final = ser_inv_df_export.sort_values(by=['Invoice Date', 'Sold To Customer Number'])
        
        # SERVED INVOICE Finalization
        
        # Step 1: Export to Excel with a blank row at the top
        temp_output = 'temp_file_output.xlsx'
        temp_file = os.path.join(import_path, temp_output)
        
        with pd.ExcelWriter(temp_file, engine='openpyxl') as writer:
            # Write to row 2 (startrow=1) to leave row 1 blank
            ser_inv_df_export_final.to_excel(writer, index=False, startrow=1)
        
        # Step 2: Load workbook and apply formatting
        wb = load_workbook(temp_file)
        ws = wb.active
        
        # Step 3: Style header (row 2)
        header_fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        header_alignment = Alignment(horizontal='center')
        
        for cell in ws[2]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Step 4: Format numeric columns with comma style
        accounting_style = NamedStyle(name="accounting_style", number_format='_(* #,##0.00_);_(* (#,##0.00);_(* "-"_);_(@_)')
        for col in range(13, 15):  # Assuming numeric columns start from column 2
            for row in range(3, ws.max_row + 1):  # Data starts from row 3
                cell = ws.cell(row=row, column=col)
                if isinstance(cell.value, (int, float)):
                    cell.style = accounting_style
        
        sum_row = ws.max_row + 1
        start_range = 3
        end_range = ws.max_row
        
        for col in range(13, 15):
            col_letter = get_column_letter(col)
            sum_cell = ws.cell(row=sum_row, column=col)
            sum_cell.value = f"=SUM({col_letter}{start_range}:{col_letter}{end_range})"
            sum_cell.font = Font(bold=True, color="FF0000")
            sum_cell.number_format = '_(* #,##0.00_);_(* (#,##0.00);_(* "-"_);_(@_)'
        
        # List the column letters or indexes you want to auto-fit
        columns_to_adjust = ['M', 'N']  # or use [2, 4] for indexes
        
        for col_id in columns_to_adjust:
            # Convert index to letter if needed
            col_letter = get_column_letter(col_id) if isinstance(col_id, int) else col_id
            max_length = 0
        
            for cell in ws[col_letter]:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
        
            ws.column_dimensions[col_letter].width = max_length + 2  # Add padding
        
        # Save changes
        full_path_ser_inv = os.path.join(export_path, ex_filename_ser_inv)
        
        ws.sheet_view.showGridLines = False
        wb.properties.author = "Joshua Ocampo"
        wb.save(full_path_ser_inv)
        # DMS Sellout Export
        
        # DMS Sellout - Invoice Preparation
        
        sellout_inv_df = df1[['Invoice Date', 'Invoice number', 'Employee ID', 'Sold To Customer Number',
                               'Sold To Customer Name', 'Item Code', 'Product/Item Description',
                                'Product UOM', 'Quantity', 'Invoice Item Type']]
        sellout_inv_df_pr = sellout_inv_df[sellout_inv_df['Invoice Item Type'] != 'ITM_SALES_TAX']
        sellout_inv_df_pr = sellout_inv_df_pr.drop('Invoice Item Type', axis=1)
        
        # DMS Sellout - Customer Returns Preparation
        
        sellout_ret_df = df2[['Customer Return Date', 'Customer Return Number', 'Sales Rep ID',
                              'Sold To Customer Number', 'Sold To Customer Name', 'Product Code',
                              'Product Description', 'UOM', 'Return/ QC Quantity', 'Customer Return Type']]
        sellout_ret_df['Return/ QC Quantity'] = -abs(sellout_ret_df['Return/ QC Quantity'])  # Add this line
        sellout_ret_df_pr = sellout_ret_df[sellout_ret_df['Customer Return Type'] != 'CUSTOMER_RETURN']
        sellout_ret_df_pr = sellout_ret_df_pr.drop('Customer Return Type', axis=1)
        
        # Filter Item Code/Product Code containing 'KNE'.
        if dummy_code == "Y":  
            sellout_inv_df_kne = sellout_inv_df_pr[sellout_inv_df_pr['Item Code'].astype(str).str.contains('KNE', na=False)]
            sellout_ret_df_kne = sellout_ret_df_pr[sellout_ret_df_pr['Product Code'].astype(str).str.contains('KNE', na=False)]
        
        # Include all items in the data frame.
        if dummy_code == "N": 
            sellout_inv_df_kne = sellout_inv_df_pr
            sellout_ret_df_kne = sellout_ret_df_pr
        
        #Rename columns
        sellout_inv_df_kne.rename(columns={
        'Invoice Date': 'Transaction Date',
        'Invoice number': 'Document Number',
        'Item Code': 'Product Code'
        }, inplace=True)
        
        sellout_ret_df_kne.rename(columns={
        'Customer Return Date': 'Transaction Date',
        'Customer Return Number': 'Document Number',
        'Sales Rep ID': 'Employee ID',
        'Product Description': 'Product/Item Description',
        'UOM': 'Product UOM',
        'Return/ QC Quantity': 'Quantity'
        }, inplace=True)
        
        
        # Join Invoice and Customer Returns
        sellout_net_df = pd.concat([sellout_inv_df_kne, sellout_ret_df_kne], ignore_index=True)
        
        sellout_net_df_final = sellout_net_df[['Document Number', 'Transaction Date', 'Employee ID',
                                               'Sold To Customer Number', 'Sold To Customer Name',
                                               'Product Code', 'Product/Item Description', 'Product UOM', 'Quantity']]
        
        # Export Net DMS Sellout
        ex_filename_sellout_net = "NET DMS Sellout.xlsx"
        full_path_sellout_net = os.path.join(export_path, ex_filename_sellout_net)
        
        export_sellout_net = sellout_net_df_final.copy()
        export_sellout_net = export_sellout_net.sort_values(by='Transaction Date')
        export_sellout_net.to_excel(full_path_sellout_net, index=False)
        
        # Set author metadata
        wb_sellout = load_workbook(full_path_sellout_net)
        wb_sellout.properties.author = "Joshua Ocampo"
        wb_sellout.save(full_path_sellout_net)
        