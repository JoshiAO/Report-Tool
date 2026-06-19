import re
with open('legacy_etl.py', 'r', encoding='utf-8') as f:
    code = f.read()

# 1. Remove tkinter and sql
code = re.sub(r'import tkinter.*?\n', '', code)
code = re.sub(r'from tkinter.*?\n', '', code)
code = re.sub(r'from sqlalchemy.*?\n', '', code)
code = re.sub(r'import pymysql.*?\n', '', code)

# 2. Remove the SQL blocks
sql_start = code.find('# Test Database Connection')
if sql_start != -1:
    code = code[:sql_start]

# 3. Replace the tkinter function definition
tkinter_func_start = code.find('def edit_missing_categories(')
tkinter_func_end = code.find('# Call editor and reload saved reference')
if tkinter_func_start != -1 and tkinter_func_end != -1:
    code = code[:tkinter_func_start] + code[tkinter_func_end:]

# 4. Indent and wrap in class method
lines = code.split('\n')
new_lines = []
for line in lines:
    if 'report_date = datetime(' in line or 'base_dir = r"D:\\COB DATA"' in line or 'category_ref_path = ' in line:
        continue
    if 'category_df, category_checking = edit_missing_categories' in line:
        new_lines.append('        category_df = await self.handle_missing_categories(category_df, all_products, category_ref_path)')
        continue
    new_lines.append('        ' + line)

final_code = '''import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, NamedStyle, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import asyncio

class LegacyETL:
    def __init__(self, settings, report_date, dummy_code, import_files, export_path, ci_path, send_progress, handle_missing_categories):
        self.settings = settings
        self.report_date = report_date
        self.dummy_code = dummy_code
        self.import_files = import_files
        self.export_path = export_path
        self.ci_path = ci_path
        self.send_progress = send_progress
        self.handle_missing_categories = handle_missing_categories

    async def execute(self):
        report_date = self.report_date
        dummy_code = self.dummy_code
        export_path = self.export_path
        import_path = os.path.dirname(self.import_files['invoice'])
        ci_path = self.ci_path
        
        category_ref_path = self.settings.reference_path_category
        field_supervisors_df = pd.read_excel(self.settings.reference_path_fs)
        week_df = pd.read_excel(self.settings.reference_path_week)
        cdam_df = pd.read_excel(self.settings.reference_path_cdam)
        gt_channel_df = pd.read_excel(self.settings.reference_path_gt_channel)
        new_customer_df = pd.read_excel(self.settings.reference_path_new_customer)
        
        # Override file names to use the uploaded ones
        im_filename_inv = self.import_files['invoice']
        im_filename_ret = self.import_files['returns']
        im_filename_cust = self.import_files['customer']
        im_filename_route = self.import_files['route']
        im_filename_pl = self.import_files['price']
        im_filename_so = self.import_files['sales_order']
        
        category_df = pd.read_excel(category_ref_path)

        full_im_path_inv = im_filename_inv
        full_im_path_ret = im_filename_ret
        full_im_path_cust = im_filename_cust
        full_im_path_route = im_filename_route
        full_im_path_pl = im_filename_pl
        full_im_path_so = im_filename_so
''' + '\n'.join(new_lines)

with open('etl_core.py', 'w', encoding='utf-8') as f:
    f.write(final_code)
