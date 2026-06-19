import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, NamedStyle, Border, Side
from openpyxl.utils import get_column_letter

def apply_header_style(ws, header_row, start_color='44546A', end_color='44546A'):
    fill = PatternFill(start_color=start_color, end_color=end_color, fill_type='solid')
    font = Font(bold=True, color='FFFFFF')
    align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    border = Border(left=Side(style=None), right=Side(style=None), top=Side(style=None), bottom=Side(style=None))
    
    ws.row_dimensions[header_row].height = 45
    for cell in ws[header_row]:
        cell.fill = fill
        cell.font = font
        cell.alignment = align
        cell.border = border

def format_accounting_columns(ws, start_col, end_col, start_row, max_row, add_sum=True):
    accounting_style = NamedStyle(name="accounting_style", number_format='_(* #,##0.00_);_(* (#,##0.00);_(* "-"_);_(@_)')
    # Try to add style to workbook, ignore if exists
    try:
        ws.parent.add_named_style(accounting_style)
    except:
        pass
        
    for col in range(start_col, end_col + 1):
        for row in range(start_row, max_row + 1):
            cell = ws.cell(row=row, column=col)
            if isinstance(cell.value, (int, float)):
                cell.style = "accounting_style"
                
    if add_sum:
        sum_row = max_row + 1
        for col in range(start_col, end_col + 1):
            col_letter = get_column_letter(col)
            sum_cell = ws.cell(row=sum_row, column=col)
            sum_cell.value = f"=SUM({col_letter}{start_row}:{col_letter}{max_row})"
            sum_cell.font = Font(bold=True, color="FF0000")
            sum_cell.style = "accounting_style"

def auto_fit_columns(ws, col_letters_or_indices, padding=2):
    for col_id in col_letters_or_indices:
        col_letter = get_column_letter(col_id) if isinstance(col_id, int) else col_id
        max_length = 0
        for cell in ws[col_letter]:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + padding
